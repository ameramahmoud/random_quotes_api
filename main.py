from fastapi import FastAPI, HTTPException, status
from fastapi.responses import Response, RedirectResponse
from fastapi.requests import Request
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from random import randint
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
from jose import JWTError, jwt
from passlib.context import CryptContext 


SECRET_KEY = "c6d620d028531f611b1cf051f6453803cbf5b7e80894d4ec352b014cd9ac110b"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60

QUOTE_IDS_TRACKER = []

templates = Jinja2Templates(directory="templates/")

app = FastAPI()

db = {
    "shebak@2023": {
    "username": "shebak@2023",
    "full_name": "shebak",
    "email": "emaill@gmail.com",
    "hashed_password": "$2b$12$vYaBSjadpnRoN6HXgFwZU.RGC/TgA9vDR4P6A7Ri.Hv4ecL6EFAuy", 
    }
}

class Token(BaseModel):
    access_token: str
    token_type: str

class TokenData(BaseModel):
    username: str or None = None

class User(BaseModel):
    username: str
    email: str or None = None
    full_name: str or None = None

class UserInDB(User):
    hashed_password: str

pwd_context = CryptContext(schemes=["bcrypt"], deprecated= "auto")

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def get_user(db, username: str):
    if username in db:
        user_data = db[username]
        return UserInDB(**user_data)
    
def authenticate_user(db, username: str, password: str):
    user = get_user(db, username)
    if not user:
        return False
    if not verify_password(password, user.hashed_password):
        return False
    
    return user

def create_access_token(data: dict, expires_delta: timedelta or None = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)

    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm= ALGORITHM)
    return encoded_jwt

async def get_current_user(token: str):
    credential_exception = HTTPException(status_code=status.HTTP_403_FORBIDDEN, 
                                        detail="You are not authorized to use this API", headers={"WWW-Authenticate": "Bearer"})
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            return credential_exception
        token_data = TokenData(username=username)
    except JWTError:
        return credential_exception
    
    user = get_user(db, username=token_data.username)
    if user is None:
        return credential_exception
    return user

# pwd = get_password_hash("shebak2023")
# print(pwd)

########################################################


@app.get("/login")
def login(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login")
async def login (response: Response, request: Request):
    form = await request.form()
    user = authenticate_user(db, form.get("username"), (form.get("password")))
    if not user:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, 
                            detail="You are not authorized to use this API", headers={"WWW-Authenticate": "Bearer"})
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires)

    response = RedirectResponse(f"/quote/random")
    response.set_cookie(key="access_token", value=f"Bearer {access_token}", httponly=True)
    return response

########################################################


def generate_random_number(data):
    """
    generates a random number within range of the recieved data
    """
    min_quote_id = data.id.min()
    max_quote_id = data.id.max()
    random_id = randint(min_quote_id, max_quote_id)
    return random_id

def generate_random_quote():
    """
    generates a random quote ID and displays it along with the quote and the author
    """

    json_quotes_data = pd.read_json("files/quotes.json")
    json_authors_data = pd.read_json("files/authors.json")

    random_id = generate_random_number(json_quotes_data)

    resulted_quote = json_quotes_data.loc[(json_quotes_data.id == random_id)]

    resulted_quote_id = str(resulted_quote.id.values[0])
    resulted_quote = str(resulted_quote.quote.values[0])
    
    QUOTE_IDS_TRACKER.append(resulted_quote_id)

    for index, qoute_ids_lists in enumerate(json_authors_data["quoteIds"]):
        if random_id in qoute_ids_lists:
            resulted_author = str(json_authors_data.iloc[index].author)
            
    if len(QUOTE_IDS_TRACKER) == 100:
        create_report()
        
    return resulted_quote_id, resulted_quote, resulted_author


def create_report():
    """
    creates an excel sheet of each quote id and its number of occurences
    """
    current_date_time = str(datetime.now().strftime("%Y_%m_%d_%H_%M_%S"))
    spreadsheet_name = "quotes_api_report_" + current_date_time + ".xlsx"

    pd.Series(QUOTE_IDS_TRACKER).value_counts().to_excel(spreadsheet_name) 
    wb = load_workbook(spreadsheet_name)
    ws = wb.active
    ws["A1"] = "Quote ID"
    ws["B1"] = "Count"
    wb.save(filename = spreadsheet_name)
    QUOTE_IDS_TRACKER.clear()


async def check_authorization_token(request: Request):
    token = request.cookies.get("access_token") 
    if token is None: 
        return {"message ": "you are not authorized to use this api"}
    
    scheme, _, parameter = token.partition(" ")
    user = await get_current_user(parameter)
    if user: 
        resulted_quote_id, resulted_quote, resulted_author = generate_random_quote()
        return templates.TemplateResponse("index.html", {"request": request, "quoteId": resulted_quote_id, "quote": resulted_quote, "author": resulted_author})
    return {"message ": "you are not authorized to use this api"}


@app.get("/quote/random")
async def get_random_quote(request: Request):
    return await check_authorization_token(request)

    
@app.post("/quote/random")
async def generate_auth_quote(request: Request):
    form = await request.form()
    return await check_authorization_token(request)